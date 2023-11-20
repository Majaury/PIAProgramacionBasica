import requests
import openpyxl

url_api = "https://pokeapi.co/api/v2/pokemon/"
url_api_gene = "https://pokeapi.co/api/v2/generation/"

def getdata(pokemon_data_url = ""):
    response = requests.get(pokemon_data_url)
    data = response.json()
    pokemon_data = {
        'pokemon_species': ""
    }
    pokemon_data['pokemon_species'] = data['pokemon_species']

    return pokemon_data


def generaciones():
    for i in range(1,10,1):
        j = str(i)
        pokemon_data_url = url_api_gene + j
    return(pokemon_data_url)

def promedio_2 (x,y):
    z = (x * 100)/y
    return z

#Experimento con excel 
try:    
    excel_datos = openpyxl.Workbook()
    hoja_2 = excel_datos.create_sheet('Hoja_2', 1)
    hoja_1 = excel_datos.active
    hoja_1.title = 'Bisqueda por generación'
    hoja_2.title = 'Bisqueda por tipo'
    a = hoja_1.cell(row = 11, column = 1, value = 'Total')
    b = hoja_1.cell(row = 1, column = 2, value = 'Generación')
    c = hoja_1.cell(row = 1, column = 3, value = 'Cantidad')
    d = hoja_1.cell(row = 1, column = 4, value = 'Porcentaje')
    Numero_pokemones_por_genera = list()
    for i in range(1,10,1):
        j = str(i)
        pokemon_data_url = (url_api_gene + j)
        data = getdata(pokemon_data_url)
        Numero_pokemones_por_genera.append(len(data['pokemon_species']))
        e = hoja_1.cell(row = i+1, column = 2, value = i)
        f = hoja_1.cell(row = i+1, column = 3, value = len(data['pokemon_species']))
    print (Numero_pokemones_por_genera)
    total = sum(Numero_pokemones_por_genera)
    h = hoja_1.cell(row = 11, column = 3, value = total)
    l = hoja_1.cell(row = 11, column = 4, value = '100%')
    for i in range (len(Numero_pokemones_por_genera)):
        w = promedio_2(Numero_pokemones_por_genera[i], total)
        g = hoja_1.cell(row = i+2, column = 4, value = w)
    excel_datos.save('Excel datos.xlsx')
    print (Numero_pokemones_por_genera)
except(requests.exceptions.RequestException, OSError) as e:
    print ('No se a podido conectar con la API, verifique su coneccion a internet')
