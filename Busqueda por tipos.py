import requests
import os
import openpyxl 

url_api_gene = "https://pokeapi.co/api/v2/generation/"
url_api_tipo = "https://pokeapi.co/api/v2/type/"


def getdata(pokemon_data_url = ""):
    response = requests.get(pokemon_data_url)
    data = response.json()
    pokemon_data = {
        'pokemon_species': ""
    }
    pokemon_data['pokemon_species'] = data['pokemon_species']

    return pokemon_data

def getdata_2(pokemon_data_url = ""):
    response = requests.get(pokemon_data_url)
    data = response.json()
    pokemon_data_2 = {
        'pokemon': ""
    }
    pokemon_data_2['pokemon'] = data['pokemon']

    return pokemon_data_2

diccionario_tipos = {
    '1':"Normal"
    '2':"Lucha" 
    '3':"Volador"
    '4':"Veneno"
    '5':"Tierra"
    '6':"Roca"
    '7':"Bicho"
    '8':"Fantasma"
    '9':"Acero"
    '10'"Fuego"
    '11':"Agua"
    '12':"Planta"
    '13':"Eléctrico" 
    '14':"Psíquico"
    '15':"Hielo"
    '16':"Dragón"
    '17':"Siniestro"
    '18':"Hada" 
} 

ubicación_archivo = 'C:\\Users\\amaur\\OneDrive\\Escritorio\\PIA Progra\\Excel datos\\Excel datos.xlsx' #Para el final hay que colocar la dirección del archivo Excel 
#Experimento con excel

if os.path.exists(ubicación_archivo):
    #Cargar datos del Excel 
    excel_datos = openpyxl.load_workbook(ubicación_archivo)
    nombre_segunda_hoja = excel_datos.sheetnames[1] 
    hoja_2 = excel_datos[nombre_segunda_hoja] 
    excel_datos.active = hoja_2
    seguro = 10
    verificación = hoja_2['A11'].value
    if verificación == seguro:
        for i in range(1, 10, 1):
            lis_gene = list() 
            for f in range(1, 19, 1):
                lis_gene.append(hoja_2.cell(row = i + 1, columna = f + 1).value)
            print(lis_gene) 
    else:
        try:
            excel_datos = openpyxl.Workbook()
            #hoja_2 = excel_datos.create_sheet('Hoja_2', 1)
            hoja_2 = excel_datos.active
            #hoja_1.title = 'Busqueda por generación'
            hoja_2.title = 'Busqueda por tipo'
            a = hoja_2.cell(row = 1, columna = 1, value = 'Generación') 
            for s in range(1,10,1):
                b = hoja_2.cell(row = s + 1, columna = 1, value = s) 
                r = str(s)
                pokemon_data_url = (url_api_gene + r)
                data = getdata(pokemon_data_url)
                tipos_generacion = list()
                for i in range (1,19,1):
                    #c = hoja_2.cell(row = 1, columna = i + 1, value = 'Tipo' i) 
                    j = str(i)
                    k = diccionario_tipos.get(j, "Tipo Desconocido") 
                    c = hoja_2.cell(row = 1, columna = i + 1, value = k) 
                    pokemon_data_url_2 = (url_api_tipo + j)
                    data_2 = getdata_2(pokemon_data_url_2)
                    cont_typo = int(0)
                    cont_1 = int(0)
                    cont_2 = int(0)
                    info_1 = data['pokemon_species']
                    info_2 = data_2['pokemon']
                    for elen in info_2:
                        cont_2 = int(0)
                        filtro_1 = info_2[cont_1]
                        filtro_2 = filtro_1['pokemon']
                        filtro_3 = filtro_2['name']
                        for elen in info_1:
                            filtro_4 = info_1[cont_2]
                            filtro_5 = filtro_4['name']
                            if filtro_3 == filtro_5:
                                cont_typo +=1
                                cont_2 +=1
                            else:
                                cont_2 +=1
                        cont_1 +=1
                    tipos_generacion.append(cont_typo)
                    d = hoja_2.cell(row = s + 1, columna = i + 1, value = cont_typo) 
                print("Distribucion de la", s, "generacion", tipos_generacion)
            excel_datos.save('Excel datos.xlsx')
        except(requests.exceptions.RequestException, OSError) as e:
            print ('No se a podido conectar con la API, verifique su conexión a internet')      
else:
    try:
        excel_datos = openpyxl.Workbook()
        hoja_2 = excel_datos.create_sheet('Hoja_2', 1)
        hoja_2 = excel_datos.active
        #hoja_1.title = 'Busqueda por generación'
        hoja_2.title = 'Busqueda por tipo'
        a = hoja_2.cell(row = 1, columna = 1, value = 'Generación') 
        for s in range(1,10,1):
            b = hoja_2.cell(row = s + 1, columna = 1, value = s) 
            r = str(s)
            pokemon_data_url = (url_api_gene + r)
            data = getdata(pokemon_data_url)
            tipos_generacion = list()
            for i in range (1,19,1):
                #c = hoja_2.cell(row = 1, columna = i + 1, value = 'Tipo' i) 
                j = str(i)
                k = diccionario_tipos.get(j, "Tipo Desconocido")
                c = hoja_2.cell(row = 1, columna = i + 1, value = k) 
                pokemon_data_url_2 = (url_api_tipo + j)
                data_2 = getdata_2(pokemon_data_url_2)
                cont_typo = int(0)
                cont_1 = int(0)
                cont_2 = int(0)
                info_1 = data['pokemon_species']
                info_2 = data_2['pokemon']
                for elen in info_2:
                    cont_2 = int(0)
                    filtro_1 = info_2[cont_1]
                    filtro_2 = filtro_1['pokemon']
                    filtro_3 = filtro_2['name']
                    for elen in info_1:
                        filtro_4 = info_1[cont_2]
                        filtro_5 = filtro_4['name']
                        if filtro_3 == filtro_5:
                            cont_typo +=1
                            cont_2 +=1
                        else:
                            cont_2 +=1
                    cont_1 +=1
                tipos_generacion.append(cont_typo)
                d = hoja_2.cell(row = s + 1, columna = i + 1, value = cont_typo) 
            print("Distribucion de la", s, "generacion", tipos_generacion)
        excel_datos.save('Excel datos.xlsx')
        excel_datos.close()
    except(requests.exceptions.RequestException, OSError) as e:
        print ('No se a podido conectar con la API, verifique su conexión a internet')

 
